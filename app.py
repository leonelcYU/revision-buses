from flask import Flask, render_template, request, jsonify, send_file
from datetime import datetime
import os
from supabase import create_client, Client
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

# Cargar variables de entorno
load_dotenv()

app = Flask(__name__)

# Configuración de Supabase
SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY")

if not SUPABASE_URL or not SUPABASE_KEY:
    raise ValueError("Las variables SUPABASE_URL y SUPABASE_KEY deben estar configuradas en el archivo .env")

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# Lista de PPUs
BUSES = [
    "LXWP57", "LXWP58", "LXWP59", "LXWP60", "LXWP61", "LXWP62", "LXWP64", "LXWP66",
    "LXWP67", "LXWP68", "LXWP69", "LXWP70", "LXWP71", "LXWP72", "LXWP73", "LXWP74",
    "LXWP75", "LXWP76", "LXWP77", "LXWP78", "LXWP79", "LXWP80", "LXWP81", "LXWP82",
    "LXWP83", "LXWP85", "LXWP86", "LXWP87", "PFTV77", "PFTW19", "PFTW20", "PFTW25",
    "PFTW26", "PFTW28", "PFTW29", "PFTW30", "PFTW31", "PFTW32", "PFTW34", "PFTW35",
    "PFTW36", "PFTW38", "PFTW39", "PFTW40", "PFTW41", "PFTW42", "PFTW44", "PFTW45",
    "PFTW46", "PFTW47", "PFTW48", "PFTW49", "PFTW50", "PFTW51", "PFTW55", "PFTW56",
    "PFTW57", "PFTW58", "PFTW59", "PFTW60", "PFTW61", "PFTW62", "PFVG75", "PFVG76",
    "PFVG77", "PFVG78", "PFVG79", "PFVG80", "PFVG82", "PFVG83", "PFVG85", "PFVG86",
    "PFVG87", "PFVG88", "PFVG89", "PFVG90", "PFVG92", "PFVG94", "PFVG95", "PFVG96",
    "PFVG97", "PFVG98", "PFVG99", "PFVH10", "PFVH11", "PFVH12", "PFVH13", "PFVH14",
    "PFVH15", "PFYC13", "PFYC14", "PFYC16", "PFYC17", "PFYC19", "PFYC20", "PFYC25",
    "PFYC26", "PFYC27", "PFYC28", "PFYC29", "PFYC31", "PFYC32", "PFYC33", "PFYC34",
    "PFYC35", "PFYC36", "PFYC37", "PFYC43", "PFYC44", "PFYC46", "PFYC49", "PFYC50",
    "PFYC53", "PFYC55", "PFYC57", "PFYC58", "PFYC60", "PFYC61", "PFYC64", "PFYC65",
    "PFYC66", "PFYC68", "PFYC69", "PFYC70", "PFYC72", "PFYC75", "PFYC76", "PFYC77",
    "PFYC79", "PFYC80", "PFYC81", "PFYC85", "PFYC88", "PFYC90", "PFZK83", "PFZK91",
    "PGBF59", "PGBY67", "PGBY72", "PGBY73", "PGBY83", "PGKP67", "PGLD42", "PGLD67",
    "PGRZ67", "PGTT95", "PGTV12", "PGWT98", "SHCV78", "SHCV83", "SHCX39", "SHCY22",
    "SHCY28", "SHXD75", "SHXD77", "SHXD79", "SHXD85", "SHXF13", "SHXF14", "SHXF29",
    "SHXF31", "SHXF84", "SHXF85", "SHXF87", "SHXF88", "SHXF90", "SHXF92", "SHXF93",
    "SHXF97", "SHXG36", "SHXG38", "SJPB21", "SJPB25", "SJPC73", "SJPD42", "SJPD44",
    "SJPD71", "SJPD72", "SJPD97", "SJPF43", "SJPF44", "SKPH70", "SKPH73", "SKPJ90",
    "SKPK18", "SKPK19", "SKPK20", "SKPK21", "SKPK22", "SKPK23", "SKPK25", "SKPK26",
    "SKPK27", "SKPK28", "SKPK31", "SKPK32", "SKPK34", "SKPK35", "SKPK37", "SKPK39",
    "SKPK40", "SKPK42", "SKPK44", "SKPK45", "SKPK62", "SKPK63", "SKPL28", "SKPL30",
    "SKPL33", "SKPL34", "SKPL36"
]

@app.route('/')
def index():
    return render_template('index.html', buses=BUSES)

@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')

@app.route('/api/search-bus', methods=['GET'])
def search_bus():
    query = request.args.get('q', '').upper()
    if not query:
        return jsonify([])
    
    matches = [bus for bus in BUSES if query in bus]
    return jsonify(matches)

@app.route('/api/submit-revision', methods=['POST'])
def submit_revision():
    try:
        data = request.json
        
        revision_data = {
            'ppu': data['ppu'],
            'fecha': data['fecha'],
            'conectividad': data['conectividad'],
            'motivo_no_conectividad': data.get('motivo_no_conectividad'),
            'norma_grafica_correcta': data['norma_grafica_correcta'],
            'disco_duro': data['disco_duro']
        }
        
        result = supabase.table('revisiones').insert(revision_data).execute()
        
        return jsonify({'success': True, 'data': result.data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/dashboard-data', methods=['GET'])
def dashboard_data():
    try:
        # Obtener todas las revisiones
        revisiones = supabase.table('revisiones').select('*').order('fecha', desc=True).execute()
        
        # Calcular estadísticas
        total = len(revisiones.data)
        conectividad_ok = sum(1 for r in revisiones.data if r['conectividad'])
        norma_grafica_ok = sum(1 for r in revisiones.data if r['norma_grafica_correcta'])
        disco_duro_ok = sum(1 for r in revisiones.data if r['disco_duro'])
        
        # Revisiones recientes
        recientes = revisiones.data[:10] if len(revisiones.data) > 10 else revisiones.data
        
        return jsonify({
            'success': True,
            'stats': {
                'total': total,
                'conectividad_ok': conectividad_ok,
                'conectividad_fail': total - conectividad_ok,
                'norma_grafica_ok': norma_grafica_ok,
                'norma_grafica_fail': total - norma_grafica_ok,
                'disco_duro_ok': disco_duro_ok,
                'disco_duro_fail': total - disco_duro_ok
            },
            'recientes': recientes
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/export-excel', methods=['GET'])
def export_excel():
    try:
        # Obtener todas las revisiones
        revisiones = supabase.table('revisiones').select('*').order('fecha', desc=True).execute()
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Revisiones de Buses"
        
        # Estilos
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='EF4444', end_color='EF4444', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Colores para estados
        green_fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
        red_fill = PatternFill(start_color='EF4444', end_color='EF4444', fill_type='solid')
        orange_fill = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')
        white_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        
        # Encabezados
        headers = ['N°', 'PPU', 'Fecha', 'Hora', 'Conectividad', 'Motivo Falla Conectividad', 
                   'Norma Gráfica', 'Disco Duro']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Ajustar anchos de columna
        column_widths = [8, 12, 14, 10, 15, 35, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Altura de la fila de encabezado
        ws.row_dimensions[1].height = 30
        
        # Datos
        for idx, rev in enumerate(revisiones.data, 1):
            row = idx + 1
            
            # Parsear fecha
            try:
                fecha_obj = datetime.fromisoformat(rev['fecha'].replace('Z', '+00:00'))
                fecha_str = fecha_obj.strftime('%d/%m/%Y')
                hora_str = fecha_obj.strftime('%H:%M')
            except:
                fecha_str = rev['fecha']
                hora_str = ''
            
            # N°
            cell = ws.cell(row=row, column=1, value=idx)
            cell.alignment = cell_alignment
            cell.border = border
            
            # PPU
            cell = ws.cell(row=row, column=2, value=rev['ppu'])
            cell.alignment = cell_alignment
            cell.border = border
            cell.font = Font(name='Arial', size=11, bold=True)
            
            # Fecha
            cell = ws.cell(row=row, column=3, value=fecha_str)
            cell.alignment = cell_alignment
            cell.border = border
            
            # Hora
            cell = ws.cell(row=row, column=4, value=hora_str)
            cell.alignment = cell_alignment
            cell.border = border
            
            # Conectividad
            cell = ws.cell(row=row, column=5, value='OK' if rev['conectividad'] else 'FALLA')
            cell.alignment = cell_alignment
            cell.border = border
            cell.fill = green_fill if rev['conectividad'] else red_fill
            cell.font = white_font
            
            # Motivo
            motivo = rev.get('motivo_no_conectividad', '') or ''
            cell = ws.cell(row=row, column=6, value=motivo)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = border
            
            # Norma Gráfica
            cell = ws.cell(row=row, column=7, value='OK' if rev['norma_grafica_correcta'] else 'FALLA')
            cell.alignment = cell_alignment
            cell.border = border
            cell.fill = green_fill if rev['norma_grafica_correcta'] else red_fill
            cell.font = white_font
            
            # Disco Duro
            cell = ws.cell(row=row, column=8, value='OK' if rev['disco_duro'] else 'FALTA')
            cell.alignment = cell_alignment
            cell.border = border
            cell.fill = green_fill if rev['disco_duro'] else orange_fill
            cell.font = white_font
            
            # Altura de fila
            ws.row_dimensions[row].height = 25
        
        # Agregar resumen al final
        summary_row = len(revisiones.data) + 3
        
        # Título del resumen
        ws.merge_cells(f'A{summary_row}:B{summary_row}')
        cell = ws.cell(row=summary_row, column=1, value='RESUMEN ESTADÍSTICO')
        cell.font = Font(name='Arial', size=13, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
        cell.alignment = header_alignment
        cell.border = border
        
        # Estadísticas
        stats = [
            ('Total Revisiones:', len(revisiones.data)),
            ('Conectividad OK:', sum(1 for r in revisiones.data if r['conectividad'])),
            ('Conectividad Falla:', sum(1 for r in revisiones.data if not r['conectividad'])),
            ('Norma Gráfica OK:', sum(1 for r in revisiones.data if r['norma_grafica_correcta'])),
            ('Norma Gráfica Falla:', sum(1 for r in revisiones.data if not r['norma_grafica_correcta'])),
            ('Disco Duro OK:', sum(1 for r in revisiones.data if r['disco_duro'])),
            ('Disco Duro Falta:', sum(1 for r in revisiones.data if not r['disco_duro'])),
        ]
        
        for i, (label, value) in enumerate(stats):
            row = summary_row + i + 1
            cell = ws.cell(row=row, column=1, value=label)
            cell.font = Font(name='Arial', size=11, bold=True)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = border
            
            cell = ws.cell(row=row, column=2, value=value)
            cell.font = Font(name='Arial', size=11, bold=True)
            cell.alignment = cell_alignment
            cell.border = border
        
        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre de archivo con fecha
        filename = f'revisiones_buses_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

if __name__ == '__main__':
    app.run(debug=True)